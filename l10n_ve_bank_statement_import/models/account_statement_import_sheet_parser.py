import logging
import re
from html.parser import HTMLParser

from odoo import _, api, models
from odoo.exceptions import UserError

try:
    from openpyxl.utils import column_index_from_string
except ImportError:
    column_index_from_string = None

_logger = logging.getLogger(__name__)

_BDV_LINE_RE = re.compile(
    r"^(?P<date>\d{2}/\d{2}/\d{4})\s+"
    r"(?P<body>.+?)\s+"
    r"(?P<operation>Nota de D[eé]bito|"
    r"Nota de Cr[eé]dito|Saldo Inicial|Saldo Final)\s+"
    r"(?P<amount>-?[\d.]+,\d{2})\s+"
    r"(?P<balance>-?[\d.]+,\d{2})\s*$"
)
_BDV_REF_RE = re.compile(r"^(?P<concept>.*?)\s+(?P<ref>\d{10,})\s*$")
_BALANCE_MARKER_CODES = {"SI", "SF"}
_BALANCE_MARKER_TEXTS = {
    "saldo inicial",
    "saldo final",
    "si",
    "sf",
}
_US_AMOUNT_RE = re.compile(r"^-?\d{1,3}(?:,\d{3})*\.\d{2}$|^-?\d+\.\d{2}$")


class _HtmlTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows = []
        self._current_row = None
        self._in_cell = False
        self._cell_parts = []

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._current_row = []
        elif tag in ("td", "th") and self._current_row is not None:
            self._in_cell = True
            self._cell_parts = []

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._in_cell:
            self._current_row.append("".join(self._cell_parts).strip())
            self._in_cell = False
        elif tag == "tr" and self._current_row is not None:
            self.rows.append(self._current_row)
            self._current_row = None

    def handle_data(self, data):
        if self._in_cell:
            self._cell_parts.append(data)


class AccountStatementImportSheetParser(models.TransientModel):
    _inherit = "account.statement.import.sheet.parser"

    def _is_openpyxl_worksheet(self, sheet):
        return hasattr(sheet, "max_row") and not hasattr(sheet, "nrows")

    def _is_html_table(self, data_file):
        sample = data_file.lstrip()[:2048].lower()
        return sample.startswith(b"<table") or b"<table" in sample

    def _is_bdv_comprobante(self, data_file):
        try:
            sample = data_file.decode("utf-8-sig")[:4096]
        except UnicodeDecodeError:
            sample = data_file.decode("latin-1")[:4096]
        sample_norm = sample.casefold()
        return "bdvenlínea" in sample_norm or "bdvenlinea" in sample_norm

    def _read_html_table_rows(self, data_file, mapping):
        try:
            text = data_file.decode(mapping.file_encoding or "utf-8")
        except UnicodeDecodeError:
            text = data_file.decode("latin-1")
        parser = _HtmlTableParser()
        parser.feed(text)
        return parser.rows

    def _is_balance_marker_value(self, value):
        if value is None:
            return False
        text = str(value).strip()
        if not text:
            return False
        if text.upper() in _BALANCE_MARKER_CODES:
            return True
        normalized = " ".join(text.casefold().split())
        return normalized in _BALANCE_MARKER_TEXTS

    def _is_opening_or_closing_balance_row(self, values, columns=None):
        if any(self._is_balance_marker_value(value) for value in values):
            return True
        if not columns:
            return False
        for column_name in (
            "debit_credit_column",
            "description_column",
            "notes_column",
        ):
            if not columns.get(column_name):
                continue
            value = self._get_values_from_column(values, columns, column_name)
            if self._is_balance_marker_value(value):
                return True
        return False

    def _read_bdv_comprobante_rows(self, data_file, mapping):
        encoding = mapping.file_encoding or "utf-8-sig"
        try:
            text = data_file.decode(encoding)
        except UnicodeDecodeError:
            text = data_file.decode("latin-1")
        rows = []
        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line or set(line) <= {"-"}:
                continue
            match = _BDV_LINE_RE.match(line)
            if not match:
                continue
            operation = match.group("operation").strip()
            body = match.group("body").strip()
            ref_match = _BDV_REF_RE.match(body)
            if ref_match:
                concept = ref_match.group("concept").strip()
                reference = ref_match.group("ref")
            else:
                concept = body
                reference = ""
            row = [
                match.group("date"),
                concept,
                reference,
                operation,
                match.group("amount"),
                match.group("balance"),
            ]
            if self._is_opening_or_closing_balance_row(row):
                continue
            rows.append(row)
        return rows

    def _parse_lines_from_rows(self, mapping, rows, currency_code):
        if not rows:
            return []
        if mapping.no_header:
            header = []
            start_idx = getattr(mapping, "csv_leading_rows_skip", 0) or 0
        else:
            header_line = mapping.header_lines_skip_count
            if header_line > 0:
                header_line -= 1
            header = rows[header_line] if header_line < len(rows) else rows[0]
            if mapping.offset_column:
                header = header[mapping.offset_column :]
            lines_skip = getattr(mapping, "lines_skip_after_header", 0) or 0
            start_idx = mapping.header_lines_skip_count + lines_skip
        columns = {
            column_name: self._get_column_indexes(header, column_name, mapping)
            for column_name in self._get_column_names()
        }
        end_idx = len(rows) - mapping.footer_lines_skip_count
        lines = []
        for row in rows[start_idx:end_idx]:
            values = row[mapping.offset_column :] if mapping.offset_column else row
            if mapping.skip_empty_lines and not any(values):
                continue
            line = self._process_row_values(values, mapping, currency_code, columns)
            if line:
                lines.append(line)
        return lines

    def _parse_lines_html(self, mapping, data_file, currency_code):
        rows = self._read_html_table_rows(data_file, mapping)
        return self._parse_lines_from_rows(mapping, rows, currency_code)

    def _parse_lines_bdv_comprobante(self, mapping, data_file, currency_code):
        rows = self._read_bdv_comprobante_rows(data_file, mapping)
        return self._parse_lines_from_rows(mapping, rows, currency_code)

    @api.model
    def parse_header(self, csv_or_xlsx, mapping):
        if (
            not isinstance(csv_or_xlsx, tuple)
            and mapping.no_header
            and getattr(mapping, "csv_leading_rows_skip", 0)
        ):
            skip = mapping.csv_leading_rows_skip
            for _ in range(skip):
                try:
                    next(csv_or_xlsx)
                except StopIteration:
                    break
        return super().parse_header(csv_or_xlsx, mapping)

    def _openpyxl_header(self, sheet, mapping):
        if mapping.no_header:
            return []
        header_line = mapping.header_lines_skip_count
        if header_line > 0:
            header_line -= 1
        row_num = header_line + 1
        header = []
        for col in range(mapping.offset_column + 1, sheet.max_column + 1):
            value = sheet.cell(row_num, col).value
            header.append(str(value).strip() if value is not None else "")
        return header

    def _parse_lines(self, mapping, data_file, currency_code):
        if self._is_html_table(data_file):
            try:
                return self._parse_lines_html(mapping, data_file, currency_code)
            except Exception as error:
                _logger.warning("No se pudo leer el archivo HTML: %s", error)
        if self._is_bdv_comprobante(data_file):
            try:
                return self._parse_lines_bdv_comprobante(
                    mapping, data_file, currency_code
                )
            except Exception as error:
                _logger.warning("No se pudo leer el comprobante BDVenlínea: %s", error)
        if data_file[:2] == b"PK":
            try:
                from io import BytesIO

                from openpyxl import load_workbook

                workbook = load_workbook(
                    BytesIO(data_file), read_only=True, data_only=True
                )
                sheet = workbook.active
                header = self._openpyxl_header(sheet, mapping)
                columns = {
                    column_name: self._get_column_indexes(header, column_name, mapping)
                    for column_name in self._get_column_names()
                }
                return self._parse_rows(
                    mapping,
                    currency_code,
                    ((workbook, sheet), data_file),
                    columns,
                )
            except Exception as error:
                _logger.warning("No se pudo leer el archivo xlsx: %s", error)
        return super()._parse_lines(mapping, data_file, currency_code)

    def _parse_column_index(self, column_str):
        if not column_str:
            return None
        try:
            if column_str.isdigit():
                return int(column_str)
            if column_index_from_string:
                return column_index_from_string(column_str.upper())
            raise UserError(
                _(
                    "openpyxl no está disponible. No se puede convertir "
                    "letra de columna a número."
                )
            )
        except (ValueError, AttributeError):
            _logger.warning("No se pudo convertir columna '%s' a índice", column_str)
            return None

    def _normalize_us_amount_for_ve_mapping(self, value, mapping):
        if not isinstance(value, str):
            return value
        if mapping.float_thousands_sep != "dot" or mapping.float_decimal_sep != "comma":
            return value
        raw = value.strip().strip('"')
        if "," in raw and raw.rfind(",") > raw.rfind("."):
            return value
        if _US_AMOUNT_RE.fullmatch(raw):
            return raw.replace(",", "").replace(".", ",")
        return value

    @api.model
    def _parse_decimal(self, value, mapping):
        value = self._normalize_us_amount_for_ve_mapping(value, mapping)
        return super()._parse_decimal(value, mapping)

    def _get_balance_from_cell(self, csv_or_xlsx, row, column, mapping):
        if not row or not column:
            return None

        column_idx = self._parse_column_index(column)
        if not column_idx:
            return None

        try:
            if isinstance(csv_or_xlsx, tuple):
                sheet = csv_or_xlsx[1]
                if row <= sheet.max_row and column_idx <= sheet.max_column:
                    cell = sheet.cell(row, column_idx)
                    value = cell.value
                    if value is not None:
                        return self._parse_decimal(value, mapping)
            else:
                return None
        except Exception as error:
            _logger.warning("Error al leer celda [%s, %s]: %s", row, column, error)
            return None
        return None

    def _strip_optional_text(self, value):
        if isinstance(value, str):
            return value.strip()
        return value

    def _strip_row_timestamp_values(self, values, columns):
        indexes = columns.get("timestamp_column") or []
        if not indexes:
            return values
        values = list(values)
        for index in indexes:
            if isinstance(index, int) and index < len(values):
                values[index] = self._strip_optional_text(values[index])
            elif not isinstance(index, int) and index in values:
                values[index] = self._strip_optional_text(values[index])
        return values

    def _strip_optional_line_strings(self, line):
        if not line:
            return line
        if "description" in line:
            line["description"] = self._strip_optional_text(line["description"])
        if "reference" in line:
            line["reference"] = self._strip_optional_text(line["reference"])
        return line

    def _process_row_values(self, values, mapping, currency_code, columns):
        if self._is_opening_or_closing_balance_row(values, columns=columns):
            return None
        values = self._strip_row_timestamp_values(values, columns)
        timestamp = self._get_values_from_column(values, columns, "timestamp_column")
        if not timestamp:
            return None
        try:
            line = super()._process_row_values(values, mapping, currency_code, columns)
        except ValueError:
            return None
        return self._strip_optional_line_strings(line)

    def _parse_rows_openpyxl(self, mapping, currency_code, sheet, columns, data_file):
        numrows = sheet.max_row
        footer_line = numrows - mapping.footer_lines_skip_count
        lines_skip_after = getattr(mapping, "lines_skip_after_header", 0) or 0
        start_row = mapping.header_lines_skip_count + 1 + lines_skip_after
        lines = []
        max_col = sheet.max_column
        for row_num in range(start_row, footer_line + 1):
            values = [
                sheet.cell(row_num, col).value
                for col in range(mapping.offset_column + 1, max_col + 1)
            ]
            if mapping.skip_empty_lines and not any(values):
                continue
            line = self._process_row_values(values, mapping, currency_code, columns)
            if line:
                lines.append(line)
        return lines

    def _parse_rows(self, mapping, currency_code, data, columns):
        csv_or_xlsx, data_file = data
        if isinstance(csv_or_xlsx, tuple):
            sheet = csv_or_xlsx[1]
            if self._is_openpyxl_worksheet(sheet):
                return self._parse_rows_openpyxl(
                    mapping, currency_code, sheet, columns, data_file
                )

        lines_skip_after = getattr(mapping, "lines_skip_after_header", 0) or 0
        if not lines_skip_after:
            return super()._parse_rows(mapping, currency_code, data, columns)

        if isinstance(csv_or_xlsx, tuple):
            numrows = csv_or_xlsx[1].max_row
        else:
            numrows = len(str(data_file.strip()).split("\n"))

        label_line = mapping.header_lines_skip_count
        footer_line = numrows - mapping.footer_lines_skip_count
        start_line = label_line + 1 + lines_skip_after

        lines = []
        if isinstance(csv_or_xlsx, tuple):
            sheet = csv_or_xlsx[1]
            max_col = sheet.max_column
            for row_num in range(start_line, footer_line + 1):
                values = [
                    sheet.cell(row_num, col).value
                    for col in range(mapping.offset_column + 1, max_col + 1)
                ]
                if mapping.skip_empty_lines and not any(values):
                    continue
                line = self._process_row_values(values, mapping, currency_code, columns)
                if line:
                    lines.append(line)
        else:
            rows = csv_or_xlsx
            for row_idx, row in enumerate(rows, label_line):
                if row_idx < start_line - 1:
                    continue
                if row_idx >= footer_line:
                    continue
                values = list(row)
                if mapping.skip_empty_lines and not any(values):
                    continue
                line = self._process_row_values(values, mapping, currency_code, columns)
                if line:
                    lines.append(line)
        return lines

    @api.model
    def parse(self, data_file, mapping, filename):
        result = super().parse(data_file, mapping, filename)
        if not result or len(result) < 3:
            return result

        currency_code, account_number, statements = result
        if not statements or not statements[0].get("transactions"):
            return result

        try:
            from io import BytesIO

            from openpyxl import load_workbook

            workbook = load_workbook(filename=BytesIO(data_file), data_only=True)
            csv_or_xlsx = (workbook, workbook.worksheets[0])
        except Exception:
            return result

        initial_balance = None
        final_balance = None

        if hasattr(mapping, "initial_balance_row") and hasattr(
            mapping, "initial_balance_column"
        ):
            if mapping.initial_balance_row and mapping.initial_balance_column:
                initial_balance = self._get_balance_from_cell(
                    csv_or_xlsx,
                    mapping.initial_balance_row,
                    mapping.initial_balance_column,
                    mapping,
                )

        if hasattr(mapping, "final_balance_row") and hasattr(
            mapping, "final_balance_column"
        ):
            if mapping.final_balance_row and mapping.final_balance_column:
                final_balance = self._get_balance_from_cell(
                    csv_or_xlsx,
                    mapping.final_balance_row,
                    mapping.final_balance_column,
                    mapping,
                )

        if initial_balance is not None or final_balance is not None:
            statement_data = statements[0]
            if initial_balance is not None:
                statement_data["balance_start"] = initial_balance
            if final_balance is not None:
                statement_data["balance_end_real"] = final_balance

        return currency_code, account_number, statements
